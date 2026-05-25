"""
Script CLI para migrar tarifas desde el Excel de comisiones a la BD.

Uso:
    cd backend
    python -m app.commands.seed_tarifas_excel
"""

import os
import re
import sys
from decimal import Decimal, InvalidOperation

import openpyxl

# Ensure project root is on path
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))

from app.database import SessionLocal
from app.models.cliente import Cliente
from app.models.comisionista import Comisionista, TipoTarifa
from app.models.producto import Producto
from app.models.cliente import Finca
from app.models.tarifa_cliente_producto import TarifaClienteProducto

EXCEL_PATH = "/Users/xstaked/Desktop/projects/sn8-projects/comisiones/Copia de COMISIONES GENERAL.xlsx"

NOMBRES_COMISIONISTAS = ["ARROYO", "CASTRO", "PINEDA", "MALAVE"]

# Normalización de nombres de productos del Excel a la BD
PRODUCTO_NORMALIZACION = {
    "PASTIILLAS": "PAST TH",
    "CALCIUM, POTASIUM, MAGNESIUM": "CALCIUM POTASIUM MAGNESIUM",
}


def normalizar_producto(nombre: str) -> str:
    return PRODUCTO_NORMALIZACION.get(nombre, nombre)


def parse_valor(celda) -> tuple[str, Decimal] | None:
    """
    Parsea el valor de una celda del Excel y retorna (tipo, valor).
    - Si es string con '$', es fijo_kg.
    - Si es número > 0, es porcentaje.
    - Si es 0 o vacío, retorna None.
    """
    if celda is None:
        return None

    if isinstance(celda, str):
        celda = celda.strip()
        if not celda or celda == "0":
            return None
        if "$" in celda:
            # Extraer número después del $
            match = re.search(r"\$?\s*(\d+(?:[.,]\d+)?)", celda)
            if match:
                try:
                    val = Decimal(match.group(1).replace(",", "."))
                    if val > 0:
                        return ("fijo_kg", val)
                except InvalidOperation:
                    pass
            return None
        # Puede ser un número como string
        try:
            val = Decimal(celda.replace(",", "."))
            if val > 0:
                return ("porcentaje", val)
        except InvalidOperation:
            pass
        return None

    if isinstance(celda, (int, float)):
        val = Decimal(str(celda))
        if val > 0:
            return ("porcentaje", val)
        return None

    return None


def obtener_o_crear_comisionistas(db):
    comisionistas = {}
    for nombre in NOMBRES_COMISIONISTAS:
        c = db.query(Comisionista).filter(Comisionista.nombre == nombre).first()
        if not c:
            c = Comisionista(nombre=nombre)
            db.add(c)
            db.flush()
            print(f"  Creado comisionista: {nombre} ({c.id})")
        else:
            print(f"  Encontrado comisionista: {nombre} ({c.id})")
        comisionistas[nombre] = c.id
    return comisionistas


def obtener_clientes(db):
    clientes = {}
    for c in db.query(Cliente).all():
        clientes[c.nombre] = c.id
    return clientes


def obtener_fincas(db):
    fincas = {}
    for f in db.query(Finca).all():
        fincas[f.nombre] = f.id
    return fincas


def obtener_productos(db):
    productos = {}
    for p in db.query(Producto).all():
        productos[p.nombre] = p.id
    return productos


def procesar_hoja(wb, nombre_hoja, cliente_default: str | None, db, comisionistas, clientes, productos, fincas):
    ws = wb[nombre_hoja]
    tarifas_creadas = 0
    tarifas_omitidas = 0

    # Leer encabezados de productos (fila 1) y comisionistas (fila 2)
    productos_por_col = {}
    comisionistas_por_col = {}

    for col in range(1, ws.max_column + 1):
        prod = ws.cell(row=1, column=col).value
        com = ws.cell(row=2, column=col).value
        if prod:
            productos_por_col[col] = normalizar_producto(str(prod).strip())
        if com and str(com).strip() in NOMBRES_COMISIONISTAS:
            comisionistas_por_col[col] = str(com).strip()

    # Determinar producto para cada columna "mirando hacia atrás"
    producto_actual_por_col = {}
    ultimo_producto = None
    for col in range(1, ws.max_column + 1):
        if col in productos_por_col:
            ultimo_producto = productos_por_col[col]
        producto_actual_por_col[col] = ultimo_producto

    # Track de combinaciones ya insertadas para evitar duplicados (idempotencia)
    combinaciones_insertadas = set()

    # Procesar filas de datos
    for row in range(3, ws.max_row + 1):
        nombre_fila = ws.cell(row=row, column=1).value
        if not nombre_fila:
            continue
        nombre_fila = str(nombre_fila).strip()

        # Ignorar filas de notas
        if nombre_fila.lower().startswith("importante"):
            continue
        if nombre_fila.lower().startswith("cuando se especifica"):
            continue

        if cliente_default:
            cliente_nombre = cliente_default
            # Para Santa Priscila, el nombre de la fila es el sector/finca
            finca_id = fincas.get(nombre_fila)
            if not finca_id:
                print(f"  ⚠ Finca no encontrada: '{nombre_fila}' (fila {row}, hoja {nombre_hoja})")
                continue
        else:
            cliente_nombre = nombre_fila
            finca_id = None

        cliente_id = clientes.get(cliente_nombre)
        if not cliente_id:
            print(f"  ⚠ Cliente no encontrado: '{cliente_nombre}' (fila {row}, hoja {nombre_hoja})")
            continue

        for col in range(2, ws.max_column + 1):
            com_nombre = comisionistas_por_col.get(col)
            if not com_nombre:
                continue

            prod_nombre = producto_actual_por_col.get(col)
            if not prod_nombre:
                continue

            producto_id = productos.get(prod_nombre)
            if not producto_id:
                print(f"  ⚠ Producto no encontrado: '{prod_nombre}' (col {col}, hoja {nombre_hoja})")
                continue

            comisionista_id = comisionistas.get(com_nombre)
            if not comisionista_id:
                continue

            celda = ws.cell(row=row, column=col).value
            parsed = parse_valor(celda)
            if not parsed:
                continue

            tipo, valor = parsed

            # Evitar duplicados por constraint única (incluye finca_id)
            clave = (comisionista_id, cliente_id, producto_id, finca_id)
            if clave in combinaciones_insertadas:
                tarifas_omitidas += 1
                continue
            combinaciones_insertadas.add(clave)

            # Crear tarifa
            tarifa = TarifaClienteProducto(
                comisionista_id=comisionista_id,
                cliente_id=cliente_id,
                producto_id=producto_id,
                finca_id=finca_id,
                tipo=TipoTarifa(tipo),
                valor=valor,
            )
            db.add(tarifa)
            tarifas_creadas += 1

    if tarifas_omitidas:
        print(f"   Tarifas omitidas por duplicado: {tarifas_omitidas}")
    return tarifas_creadas


def main():
    print("=" * 60)
    print("Seed de tarifas desde Excel")
    print("=" * 60)

    if not os.path.exists(EXCEL_PATH):
        print(f"ERROR: No se encontró el archivo Excel: {EXCEL_PATH}")
        sys.exit(1)

    db = SessionLocal()
    try:
        # Idempotencia: limpiar tarifas existentes
        print("\n1. Eliminando tarifas existentes...")
        count_del = db.query(TarifaClienteProducto).delete()
        db.commit()
        print(f"   Eliminadas {count_del} tarifas.")

        # Obtener/crear comisionistas
        print("\n2. Verificando comisionistas...")
        comisionistas = obtener_o_crear_comisionistas(db)
        db.commit()

        # Obtener clientes y productos
        print("\n3. Cargando clientes, fincas y productos...")
        clientes = obtener_clientes(db)
        fincas = obtener_fincas(db)
        productos = obtener_productos(db)
        print(f"   Clientes: {len(clientes)}")
        print(f"   Fincas: {len(fincas)}")
        print(f"   Productos: {len(productos)}")

        # Abrir Excel
        print(f"\n4. Leyendo Excel: {EXCEL_PATH}")
        wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)

        total_tarifas = 0

        # Procesar SANTA PRISCILA
        if "SANTA PRISCILA" in wb.sheetnames:
            print("\n5. Procesando hoja 'SANTA PRISCILA'...")
            n = procesar_hoja(
                wb, "SANTA PRISCILA", "Santa Priscila",
                db, comisionistas, clientes, productos, fincas
            )
            db.commit()
            print(f"   Tarifas creadas: {n}")
            total_tarifas += n
        else:
            print("   Hoja 'SANTA PRISCILA' no encontrada.")

        # Procesar OTRAS EMPRESAS
        if "OTRAS EMPRESAS" in wb.sheetnames:
            print("\n6. Procesando hoja 'OTRAS EMPRESAS'...")
            n = procesar_hoja(
                wb, "OTRAS EMPRESAS", None,
                db, comisionistas, clientes, productos, fincas
            )
            db.commit()
            print(f"   Tarifas creadas: {n}")
            total_tarifas += n
        else:
            print("   Hoja 'OTRAS EMPRESAS' no encontrada.")

        print("\n" + "=" * 60)
        print(f"TOTAL DE TARIFAS CREADAS: {total_tarifas}")
        print("=" * 60)

    except Exception as exc:
        db.rollback()
        print(f"\nERROR: {exc}")
        raise
    finally:
        db.close()


if __name__ == "__main__":
    main()
