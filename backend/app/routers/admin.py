from fastapi import APIRouter, Depends, HTTPException, status
import openpyxl
from sqlalchemy.orm import Session

from app.database import get_db
from app.models.user import User
from app.dependencies import get_current_superuser
from app.commands.seed_catalogos import seed_clientes, seed_productos, truncate_catalogos
from app.commands.seed_demo import truncate_all
from app.commands.seed_tarifas_excel import (
    EXCEL_PATH,
    obtener_clientes,
    obtener_fincas,
    obtener_o_crear_comisionistas,
    obtener_productos,
    procesar_hoja,
)

router = APIRouter()


def cargar_datos_reales(db: Session) -> int:
    truncate_all(db)
    truncate_catalogos(db)
    seed_clientes(db)
    seed_productos(db)

    comisionistas = obtener_o_crear_comisionistas(db)
    db.commit()

    clientes = obtener_clientes(db)
    fincas = obtener_fincas(db)
    productos = obtener_productos(db)
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)

    total_tarifas = 0
    if "SANTA PRISCILA" in wb.sheetnames:
        total_tarifas += procesar_hoja(
            wb,
            "SANTA PRISCILA",
            "Santa Priscila",
            db,
            comisionistas,
            clientes,
            productos,
            fincas,
        )
        db.commit()

    if "OTRAS EMPRESAS" in wb.sheetnames:
        total_tarifas += procesar_hoja(
            wb,
            "OTRAS EMPRESAS",
            None,
            db,
            comisionistas,
            clientes,
            productos,
            fincas,
        )
        db.commit()

    return total_tarifas


@router.post("/seed-real")
def seed_real(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_superuser),
):
    try:
        total_tarifas = cargar_datos_reales(db)
        return {
            "detail": "Datos reales cargados correctamente",
            "total_tarifas": total_tarifas,
        }
    except Exception as exc:
        db.rollback()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error al cargar datos reales: {exc}",
        ) from exc


@router.post("/seed-demo")
def seed_demo(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_superuser),
):
    return seed_real(db, current_user)
