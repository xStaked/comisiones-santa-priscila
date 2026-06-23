#!/usr/bin/env python3
"""
Compara las tarifas del Excel (fuente de verdad) contra las del sistema.
SOLO LECTURA: no modifica nada.

Detecta por qué un comisionista comisiona en sectores donde el Excel dice 0:
  - EXTRA   : tarifa en el sistema que el Excel NO tiene (sector que debería ser 0).
  - FALTA   : tarifa en el Excel que el sistema NO tiene.
  - DIFIERE : mismo (comisionista, cliente, producto, finca) con tipo/valor distinto.
  - GLOBALES: comisionistas con tarifas globales (Tarifa) que ignoran el sector;
              si además no tienen específicas activas, comisionan TODO sector.

Uso (en el servidor con la BD de producción configurada en .env):
    cd backend
    python compare_tarifas_excel.py                 # todos los comisionistas
    python compare_tarifas_excel.py ARROYO          # filtra por comisionista
    EXCEL_PATH=/ruta/al.xlsx python compare_tarifas_excel.py
"""
import os
import sys
from decimal import Decimal

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

import openpyxl
from app.database import SessionLocal
from app.models.cliente import Cliente, Finca
from app.models.comisionista import Comisionista, Tarifa
from app.models.producto import Producto
from app.models.tarifa_cliente_producto import TarifaClienteProducto
from app.commands.seed_tarifas_excel import (
    EXCEL_PATH,
    NOMBRES_COMISIONISTAS,
    normalizar_producto,
    parse_valor,
)

# (cliente_nombre, finca_nombre) por hoja. None = finca por nombre de fila.
HOJAS = {"SANTA PRISCILA": "Santa Priscila", "OTRAS EMPRESAS": None}


def construir_esperado(wb):
    """Reconstruye lo que un seed fiel produciría: clave -> (tipo, valor).

    Espeja exactamente procesar_hoja() de seed_tarifas_excel.py.
    clave = (comisionista, cliente, producto, finca|None)
    """
    esperado = {}
    for hoja, cliente_default in HOJAS.items():
        if hoja not in wb.sheetnames:
            continue
        ws = wb[hoja]
        productos_por_col, comis_por_col = {}, {}
        for col in range(1, ws.max_column + 1):
            prod = ws.cell(row=1, column=col).value
            com = ws.cell(row=2, column=col).value
            if prod:
                productos_por_col[col] = normalizar_producto(str(prod).strip())
            if com and str(com).strip() in NOMBRES_COMISIONISTAS:
                comis_por_col[col] = str(com).strip()
        # producto "mirando hacia atrás"
        prod_actual, ultimo = {}, None
        for col in range(1, ws.max_column + 1):
            if col in productos_por_col:
                ultimo = productos_por_col[col]
            prod_actual[col] = ultimo

        vistos = set()
        for row in range(3, ws.max_row + 1):
            nombre_fila = ws.cell(row=row, column=1).value
            if not nombre_fila:
                continue
            nombre_fila = str(nombre_fila).strip()
            low = nombre_fila.lower()
            if low.startswith("importante") or low.startswith("cuando se especifica"):
                continue
            if cliente_default:
                cliente_nombre, finca_nombre = cliente_default, nombre_fila
            else:
                cliente_nombre, finca_nombre = nombre_fila, None
            for col in range(2, ws.max_column + 1):
                com_nombre = comis_por_col.get(col)
                prod_nombre = prod_actual.get(col)
                if not com_nombre or not prod_nombre:
                    continue
                parsed = parse_valor(ws.cell(row=row, column=col).value)
                if not parsed:
                    continue
                clave = (com_nombre, cliente_nombre, prod_nombre, finca_nombre)
                if clave in vistos:
                    continue
                vistos.add(clave)
                esperado[clave] = parsed  # (tipo, valor)
    return esperado


def construir_sistema(db):
    """clave -> (tipo, valor) de las tarifas activas del sistema."""
    sistema = {}
    q = (
        db.query(TarifaClienteProducto, Comisionista.nombre, Cliente.nombre,
                 Producto.nombre, Finca.nombre)
        .join(Comisionista, TarifaClienteProducto.comisionista_id == Comisionista.id)
        .join(Cliente, TarifaClienteProducto.cliente_id == Cliente.id)
        .join(Producto, TarifaClienteProducto.producto_id == Producto.id)
        .outerjoin(Finca, TarifaClienteProducto.finca_id == Finca.id)
        .filter(TarifaClienteProducto.activo.is_(True))
    )
    for tar, com, cli, prod, fin in q.all():
        clave = (com, cli, prod, fin)
        sistema[clave] = (tar.tipo.value, Decimal(str(tar.valor)))
    return sistema


def fmt(clave):
    com, cli, prod, fin = clave
    return f"{com:8} | {cli:18} | {prod:22} | finca={fin or '(sin finca)'}"


def main():
    filtro = sys.argv[1].upper() if len(sys.argv) > 1 else None
    if not os.path.exists(EXCEL_PATH):
        print(f"ERROR: Excel no encontrado: {EXCEL_PATH}")
        sys.exit(1)

    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    esperado = construir_esperado(wb)

    db = SessionLocal()
    try:
        sistema = construir_sistema(db)

        def keep(clave):
            return filtro is None or clave[0].upper() == filtro

        esp = {k: v for k, v in esperado.items() if keep(k)}
        sis = {k: v for k, v in sistema.items() if keep(k)}

        extra = sorted(set(sis) - set(esp))
        falta = sorted(set(esp) - set(sis))
        comun = sorted(set(esp) & set(sis))
        difiere = [(k, esp[k], sis[k]) for k in comun if esp[k] != sis[k]]

        print("=" * 90)
        print(f"COMPARACIÓN EXCEL (fuente) vs SISTEMA   filtro={filtro or 'TODOS'}")
        print(f"Excel: {EXCEL_PATH}")
        print("=" * 90)
        print(f"Esperadas (Excel): {len(esp)} | En sistema: {len(sis)} | "
              f"Coinciden: {len(comun)-len(difiere)}")

        print(f"\n### EXTRA en sistema ({len(extra)}) — comisiona donde Excel dice 0:")
        for k in extra:
            print(f"  + {fmt(k)} | sistema={sis[k][0]} {sis[k][1]}")

        print(f"\n### DIFIERE valor/tipo ({len(difiere)}):")
        for k, e, s in difiere:
            print(f"  ~ {fmt(k)} | excel={e[0]} {e[1]}  ->  sistema={s[0]} {s[1]}")

        print(f"\n### FALTA en sistema ({len(falta)}):")
        for k in falta:
            print(f"  - {fmt(k)} | excel={esp[k][0]} {esp[k][1]}")

        # Fuga por tarifas globales (ignoran sector)
        print("\n### TARIFAS GLOBALES (Tarifa) — ignoran sector:")
        coms = db.query(Comisionista).all()
        for c in coms:
            if filtro and c.nombre.upper() != filtro:
                continue
            globales = list(c.tarifas)
            n_esp = (
                db.query(TarifaClienteProducto)
                .filter(TarifaClienteProducto.comisionista_id == c.id,
                        TarifaClienteProducto.activo.is_(True))
                .count()
            )
            if globales:
                fuga = " <-- ¡FUGA! sin específicas activas: globales aplican a TODO sector" if n_esp == 0 else ""
                detalle = ", ".join(f"{t.tipo.value} {t.valor}" for t in globales)
                print(f"  {c.nombre:10} globales=[{detalle}] | específicas_activas={n_esp}{fuga}")
    finally:
        db.close()


if __name__ == "__main__":
    main()
