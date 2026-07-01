#!/usr/bin/env python3
"""Solo-lectura: compara TIPO de tarifa (porcentaje vs fijo_kg) Excel vs prod
para los comisionistas internos, usando la MISMA normalización del backend.

Excel: '$' => fijo_kg ; numero plano => porcentaje  (igual que parse_valor).
Uso: DATABASE_URL=... EXCEL_PATH=... .venv/bin/python diag_tipo_excel_vs_prod.py
"""
import os, sys
from decimal import Decimal
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

import openpyxl
from app.database import SessionLocal
from app.models.cliente import Cliente, Finca
from app.models.comisionista import Comisionista
from app.models.producto import Producto
from app.models.tarifa_cliente_producto import TarifaClienteProducto
from app.services.catalog_normalization import (
    _normalizar_texto, normalizar_nombre_finca, normalizar_nombre_producto,
)
from app.commands.seed_tarifas_excel import (
    EXCEL_PATH, NOMBRES_COMISIONISTAS, normalizar_producto, parse_valor,
)

HOJAS = {"SANTA PRISCILA": "Santa Priscila", "OTRAS EMPRESAS": None}


def construir_esperado(wb):
    esperado = {}
    for hoja, cliente_default in HOJAS.items():
        if hoja not in wb.sheetnames:
            continue
        ws = wb[hoja]
        productos_por_col, comis_por_col = {}, {}
        for col in range(1, ws.max_column + 1):
            prod = ws.cell(row=1, column=col).value
            com = ws.cell(row=2, column=col).value
            if prod:
                productos_por_col[col] = normalizar_producto(str(prod).strip())
            if com and str(com).strip() in NOMBRES_COMISIONISTAS:
                comis_por_col[col] = str(com).strip()
        prod_actual, ultimo = {}, None
        for col in range(1, ws.max_column + 1):
            if col in productos_por_col:
                ultimo = productos_por_col[col]
            prod_actual[col] = ultimo
        for row in range(3, ws.max_row + 1):
            fila = ws.cell(row=row, column=1).value
            if not fila:
                continue
            fila = str(fila).strip()
            low = fila.lower()
            if low.startswith("importante") or low.startswith("cuando se especifica"):
                continue
            if cliente_default:
                cli_n = _normalizar_texto(cliente_default)
                fin_n = normalizar_nombre_finca(fila)
            else:
                cli_n = _normalizar_texto(fila)
                fin_n = None
            for col in range(2, ws.max_column + 1):
                com = comis_por_col.get(col)
                prod = prod_actual.get(col)
                if not com or not prod:
                    continue
                parsed = parse_valor(ws.cell(row=row, column=col).value)
                if not parsed:
                    continue
                clave = (com.upper(), cli_n, normalizar_nombre_producto(prod), fin_n)
                if clave not in esperado:
                    esperado[clave] = parsed
    return esperado


def construir_sistema(db):
    sistema = {}
    q = (
        db.query(TarifaClienteProducto, Comisionista.nombre, Cliente.nombre,
                 Producto.nombre, Finca.nombre)
        .join(Comisionista, TarifaClienteProducto.comisionista_id == Comisionista.id)
        .join(Cliente, TarifaClienteProducto.cliente_id == Cliente.id)
        .join(Producto, TarifaClienteProducto.producto_id == Producto.id)
        .outerjoin(Finca, TarifaClienteProducto.finca_id == Finca.id)
        .filter(TarifaClienteProducto.activo.is_(True))
    )
    for tar, com, cli, prod, fin in q.all():
        clave = (com.upper(), _normalizar_texto(cli),
                 normalizar_nombre_producto(prod),
                 normalizar_nombre_finca(fin) if fin else None)
        sistema[clave] = (tar.tipo.value, Decimal(str(tar.valor)))
    return sistema


def main():
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    esperado = construir_esperado(wb)
    db = SessionLocal()
    try:
        sistema = construir_sistema(db)
        for com in NOMBRES_COMISIONISTAS:
            esp = {k: v for k, v in esperado.items() if k[0] == com}
            sis = {k: v for k, v in sistema.items() if k[0] == com}
            comun = set(esp) & set(sis)
            tipo_mismatch = [(k, esp[k], sis[k]) for k in comun if esp[k][0] != sis[k][0]]
            val_mismatch = [(k, esp[k], sis[k]) for k in comun
                            if esp[k][0] == sis[k][0] and esp[k][1] != sis[k][1]]
            print(f"\n=== {com} ===  Excel={len(esp)} prod={len(sis)} "
                  f"coinciden_clave={len(comun)} | tipo_dif={len(tipo_mismatch)} "
                  f"valor_dif={len(val_mismatch)}")
            # resumen de tipos en prod
            from collections import Counter
            ct_p = Counter(v[0] for v in sis.values())
            ct_e = Counter(v[0] for v in esp.values())
            print(f"    tipos Excel: {dict(ct_e)}  | tipos prod: {dict(ct_p)}")
            for k, e, s in tipo_mismatch[:6]:
                print(f"    TIPO  {k[1][:16]:16} {k[2][:20]:20} fin={str(k[3])[:12]:12}"
                      f" excel={e[0]} {e[1]} -> prod={s[0]} {s[1]}")
            if len(tipo_mismatch) > 6:
                print(f"    ... y {len(tipo_mismatch)-6} más con tipo distinto")
            for k, e, s in val_mismatch[:10]:
                print(f"    VALOR {k[1][:16]:16} {k[2][:20]:20} fin={str(k[3])[:12]:12}"
                      f" excel={e[1]} -> prod={s[1]}")
            if len(val_mismatch) > 10:
                print(f"    ... y {len(val_mismatch)-10} más con valor distinto")
    finally:
        db.close()


if __name__ == "__main__":
    main()
